import pandas as pd

def read_excel_fast(*args, **kwargs):
    try:
        return pd.read_excel(*args, engine='calamine', **kwargs)
    except Exception as e:
        # Fallback to default engine (openpyxl) in case of any calamine incompatibilities
        return pd.read_excel(*args, **kwargs)

import datetime
import openpyxl
from excel_generator import ExcelGenerator
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
import os
import zipfile
import tempfile
import json
import re
import traceback
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_unique_filename(output_folder, base_name, extension):
    clean_base = re.sub(r'[\\/*?:\u0022<>|]', "-", base_name)
    path = os.path.join(output_folder, f"{clean_base}.{extension}")
    if not os.path.exists(path):
        return path
    counter = 1
    while True:
        path = os.path.join(output_folder, f"{clean_base} ({counter}).{extension}")
        if not os.path.exists(path):
            return path
        counter += 1

class ReportProcessor:
    def __init__(self, sap_file, loginext_file, output_folder, report_type="E-Com", remaining_file=None):
        self.sap_file = sap_file
        self.loginext_file = loginext_file
        self.output_folder = output_folder
        self.report_type = report_type
        self.remaining_file = remaining_file
        
    def process(self):
        try:

            if self.sap_file.lower().endswith('.zip'):
                ext_dir = tempfile.mkdtemp()
                with zipfile.ZipFile(self.sap_file, 'r') as zf:
                    zf.extractall(ext_dir)
                for r, _, fs in os.walk(ext_dir):
                    for ff in fs:
                        if not ff.startswith('~') and not ff.startswith('.') and ff.lower().endswith(('.csv', '.xlsx', '.xls', '.pdf')):
                            self.sap_file = os.path.join(r, ff)
                            break

            def normalize_desc(d):
                return re.sub(r'\\s+', ' ', str(d)).strip().lower()
            def clean_alphanumeric(d):
                return re.sub(r'[^a-z0-9]', '', normalize_desc(d))
            def is_desc_header(h):
                h_lower = str(h).lower()
                if any(x in h_lower for x in ["external id", "model number", "model num", "sku", "code"]):
                    return False
                return any(x in h_lower for x in ["title", "desc", "item", "product"])
            def is_qty_header(h):
                h_lower = str(h).lower()
                if "accepted" in h_lower or "received" in h_lower:
                    return False
                return any(x in h_lower for x in ["qty", "quantity"])
            def is_price_header(h):
                h_lower = str(h).lower()
                return any(x in h_lower for x in ["price", "rate", "amount"])
            # 1. Read files
            sap_df = read_excel_fast(self.sap_file)
            sap_df["Is Remaining"] = False
            
            # Find the reference column dynamically from SAP first
            ref_col = None
            for col in sap_df.columns:
                if str(col).lower() in ["your reference", "customer reference", "customer reference number", "customer ref"]:
                    ref_col = col
                    break
                    
            if not ref_col:
                return False, f"Error: Reference column not found in SAP file. Available columns: {', '.join(sap_df.columns.tolist()[:10])}"

            if self.remaining_file and os.path.exists(self.remaining_file):
                rem_df = read_excel_fast(self.remaining_file)
                rem_df["Is Remaining"] = True
                
                # If rem_df is a previous report, it likely has "Order Number" instead of ref_col.
                # We must rename it back to ref_col so pd.concat aligns them perfectly.
                if ref_col not in rem_df.columns:
                    for c in rem_df.columns:
                        if str(c).lower() in ["order number", "order no.", "your reference", "customer reference"]:
                            rem_df.rename(columns={c: ref_col}, inplace=True)
                            break
                            
                sap_df = pd.concat([sap_df, rem_df], ignore_index=True)
                
            loginext_df = pd.read_csv(self.loginext_file)
            if "Order No." not in loginext_df.columns:
                return False, "Error: 'Order No.' column not found in Loginext file."
            
            required_loginext_cols = ["Order No.", "End Date", "Order Status", "Delivery Associate"]
            
            missing_cols = [col for col in required_loginext_cols if col not in loginext_df.columns]
            if missing_cols:
                return False, f"Error: Missing columns in Loginext file: {', '.join(missing_cols)}"
            
            # 2. Clean text in matching columns
            # Remove '#', spaces, '=', '(', ')', '"' and leading zeros
            sap_df[ref_col] = sap_df[ref_col].astype(str).str.replace(r'[="()#\s]', '', regex=True).str.lstrip('0').str.upper()
            
            # If the SAP file already has a native "Order Number" column, it will conflict when we 
            # rename "Your Reference" to "Order Number". Drop the original one first.
            if "Order Number" in sap_df.columns and ref_col != "Order Number":
                sap_df = sap_df.drop(columns=["Order Number"])
                
            sap_df.rename(columns={ref_col: "Order Number"}, inplace=True)
            
            # For E-Com, drop 'Sales man Name' because we use Loginext's Delivery Associate for Sales Man.
            # For Montana, KEEP the 'Sales man Name' from SAP.
            if self.report_type == "E-Com" and "Sales man Name" in sap_df.columns:
                sap_df = sap_df.drop(columns=["Sales man Name"])
                
            # ALWAYS drop 'Order Status' from SAP if it exists, so Loginext's 'Order Status' merges cleanly
            # without creating _x and _y columns!
            if "Order Status" in sap_df.columns:
                sap_df = sap_df.drop(columns=["Order Status"])
            
            # Clean loginext Order No. aggressively
            loginext_df["Order No."] = loginext_df["Order No."].astype(str).str.replace(r'[="()#\s]', '', regex=True).str.lstrip('0').str.upper()
            
            # Clean loginext End Date and Delivery Associate from =("...") formulas
            for col in ["End Date", "Order Status", "Delivery Associate"]:
                if col in loginext_df.columns:
                    # Strip leading/trailing = ( ) " 
                    loginext_df[col] = loginext_df[col].astype(str).str.replace(r'^[="()]+|[="()]+$', '', regex=True)
                    # Convert string 'nan' back to empty
                    loginext_df[col] = loginext_df[col].replace('nan', '')
            
            # 3. Filter Loginext columns
            loginext_subset = loginext_df[required_loginext_cols].copy()
            
            # 4. Perform Left Join
            merged_df = pd.merge(
                sap_df, 
                loginext_subset, 
                how="left", 
                left_on="Order Number", 
                right_on="Order No."
            )
            
            # Drop the duplicate key column from right if it's there
            if "Order No." in merged_df.columns and "Order Number" in merged_df.columns:
                merged_df = merged_df.drop(columns=["Order No."])
                
            # Overwrite SAP columns with Loginext data as requested
            # Rename 'Customer Ref. Date' to 'Date' and keep only the date part from 'End Date'
            def clean_date(d):
                s = str(d).strip()
                return s.split(" ")[0] if s and s.lower() != "nan" else ""
            
            if "End Date" in merged_df.columns:
                merged_df["Date"] = merged_df["End Date"].apply(clean_date)
            else:
                merged_df["Date"] = ""
            
            # For E-Com, overwrite Sales Man with Delivery Associate
            # For Montana, KEEP the Sales Man from SAP
            if self.report_type == "E-Com" and "Delivery Associate" in merged_df.columns:
                merged_df["Sales Man"] = merged_df["Delivery Associate"]
            
            if "Customer Ref. Date" in merged_df.columns:
                merged_df = merged_df.drop(columns=["Customer Ref. Date"])
            
            # Drop the extra columns that were brought in from Loginext
            merged_df = merged_df.drop(columns=["End Date", "Delivery Associate"], errors='ignore')
                
            # Reorder columns to place 'Order Status' where 'Sales man Name' used to be
            desired_order = [
                "Date", "Order Number", "Order Status", "Sales Order", 
                "Billing Document Number", "Sales Man", "Driver Name", "Customer", "Name", 
                "Sales Header Value", "Terms of Payment", "Delivery Number", "Is Remaining"
            ]
            # Ensure we only include columns that actually exist in the merged dataframe
            final_cols = [col for col in desired_order if col in merged_df.columns]
            merged_df = merged_df[final_cols]
            
            # Explicitly drop the salesman column for Montana if it inadvertently matched
            if self.report_type == "Montana":
                for drop_col in ["Sales Man", "Sales man Name", "Sales man"]:
                    if drop_col in merged_df.columns:
                        merged_df = merged_df.drop(columns=[drop_col])

            
            # Convert Order Number to numeric safely without erasing non-numeric strings
            if "Order Number" in merged_df.columns:
                original_order_numbers = merged_df["Order Number"].copy()
                numeric_order_numbers = pd.to_numeric(merged_df["Order Number"], errors='coerce')
                # Keep numeric where possible, otherwise keep original string
                merged_df["Order Number"] = numeric_order_numbers.fillna(original_order_numbers)
            
            # Find the report date from the Date column
            valid_dates = merged_df[merged_df["Date"] != ""]["Date"]
            report_date = str(valid_dates.iloc[0]) if not valid_dates.empty else datetime.datetime.now().strftime("%d-%m-%Y")
            
            # Sanitize report_date for file paths (replace slashes with dashes)
            safe_report_date = report_date.replace('/', '-').replace('\\', '-')
                
            # Handle missing values safely (replace NaN with empty string)
            # to_excel handles NaN automatically, so we don't need to force fillna("") which breaks on float columns in Pandas 3.0.
            today_str = datetime.datetime.now().strftime("%d-%m-%Y")
            report_name_prefix = "E-Com Cash Sale Report" if self.report_type == "E-Com" else "Montana Cash Sale Report"
            
            base_filename = f"{report_name_prefix} - {safe_report_date}"
            output_path = get_unique_filename(self.output_folder, base_filename, "xlsx")
            
            # Guarantee the directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save data starting at row 3 (index 2)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, startrow=2, sheet_name="Report")
                
            # 6. Formatting with OpenPyXL
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Report"]
            
            # Hide gridlines to give a clean white background
            ws.sheet_view.showGridLines = False
            
            # Freeze panes at row 4
            ws.freeze_panes = 'A4'
            
            # Find the 'Terms of Payment' column index for the merge boundary
            pt_col_idx = ws.max_column
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=c).value == "Terms of Payment":
                    pt_col_idx = c
                    break
            
            pt_col_letter = openpyxl.utils.get_column_letter(pt_col_idx)
            
            # Add Title at top
            ws.row_dimensions[1].height = 27
            ws['A1'] = report_name_prefix
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f"A1:{pt_col_letter}1")
            
            # Add Report Date
            ws.row_dimensions[2].height = 27
            ws['A2'] = str(report_date)
            ws['A2'].font = Font(size=16, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f"A2:{pt_col_letter}2")
            
            header_row_idx = 3
            max_col = ws.max_column
            max_row = ws.max_row
            
            alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Montana report header a little bit dark
            h_color = "B4C6E7" if self.report_type == "Montana" else "DDEBF7"
            header_fill = PatternFill(start_color=h_color, end_color=h_color, fill_type="solid")
            
            for row in ws.iter_rows(min_row=header_row_idx, max_row=max_row, min_col=1, max_col=max_col):
                # Set row height to 25
                ws.row_dimensions[row[0].row].height = 25
                for cell in row:
                    cell.alignment = alignment_style
                    if cell.row == header_row_idx:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
            
            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                # Use a cell from row 3 (index 2) to get column_letter safely, avoiding merged header cells
                try:
                    column = col[2].column_letter
                except AttributeError:
                    continue
                
                # Check widths starting from header row
                for cell in col:
                    if cell.row >= header_row_idx:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                if adjusted_width > 40:
                    adjusted_width = 40
                ws.column_dimensions[column].width = adjusted_width
            
            # --- New Logic for Total Row and Summary Table ---
            N = ws.max_row
            
            # Find the Is Remaining column index and highlight rows
            is_rem_col_idx = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=col_idx).value == "Is Remaining":
                    is_rem_col_idx = col_idx
                    break
                    
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            if is_rem_col_idx:
                for row_idx in range(4, N + 1):
                    if ws.cell(row=row_idx, column=is_rem_col_idx).value == True:
                        for col_idx in range(1, is_rem_col_idx): # Don't highlight the marker column itself
                            ws.cell(row=row_idx, column=col_idx).fill = highlight_fill
                
                # Delete the Is Remaining column so it doesn't show in the final report
                ws.delete_cols(is_rem_col_idx)
            
            # The max column may have changed after deletion
            N = ws.max_row
            total_row = N + 6  # Leave 5 empty rows before TOTAL
            
            # Set Total row size
            ws.row_dimensions[total_row].height = 25
            
            # Set height for the 5 empty rows
            for r in range(N + 1, total_row):
                ws.row_dimensions[r].height = 25
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Apply AutoFilter from Row 3 (Headers) down to N
            final_max_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A3:{final_max_col_letter}{N}"
            
            # Apply borders from row 1 to total_row across columns
            final_max_col = ws.max_column
            for row_idx in range(1, total_row + 1):
                # Only draw borders up to Terms of Payment for the first two rows to keep it clean
                border_end_col = pt_col_idx if row_idx in (1, 2) else final_max_col
                for col_idx in range(1, border_end_col + 1):
                    ws.cell(row=row_idx, column=col_idx).border = thin_border
            
            # Bold and size 12 for specific columns: Customer, Name, Sales Header Value, Terms of Payment
            bold_font_12 = Font(size=12, bold=True)
            target_cols = ["Customer", "Name", "Sales Header Value", "Terms of Payment"]
            target_indices = []
            for c in range(1, final_max_col + 1):
                if ws.cell(row=3, column=c).value in target_cols:
                    target_indices.append(c)
            
            for r in range(4, N + 1):
                for c in target_indices:
                    ws.cell(row=r, column=c).font = bold_font_12
                    
            # We removed the Excel Table (ws.add_table) feature so you can freely add rows manually 
            # without Excel blocking you.
            
            # Find the Sales Header Value column to calculate total
            shv_col_idx = None
            for c in range(1, final_max_col + 1):
                if ws.cell(row=3, column=c).value == "Sales Header Value":
                    shv_col_idx = c
                    break
            
            if shv_col_idx and shv_col_idx > 1:
                # Merge columns before the Sales Header Value for the "TOTAL" label
                ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=shv_col_idx - 1)
                ws.cell(row=total_row, column=1, value="TOTAL")
                ws.cell(row=total_row, column=1).font = Font(size=13, bold=True)
                ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                # Merge Sales Header Value and any remaining columns for the Subtotal formula
                if final_max_col > shv_col_idx:
                    ws.merge_cells(start_row=total_row, start_column=shv_col_idx, end_row=total_row, end_column=final_max_col)
                shv_letter = openpyxl.utils.get_column_letter(shv_col_idx)
                ws.cell(row=total_row, column=shv_col_idx, value=f"=SUBTOTAL(9,{shv_letter}4:{shv_letter}{total_row-1})")
                ws.cell(row=total_row, column=shv_col_idx).font = Font(size=13, bold=True)
                ws.cell(row=total_row, column=shv_col_idx).alignment = Alignment(horizontal='center', vertical='center')
            
            # Leave two rows and add Summary Table
            summary_start = total_row + 3
            labels = [("CASH", "CASH"), ("POS", "POS"), ("PREPAID", "PAYT"), ("ONLINE", "ONLINE"), ("CREDIT", "CREDIT")]
            
            for i, (label, term) in enumerate(labels):
                row = summary_start + i
                ws.row_dimensions[row].height = 25
                ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
                
                for c in [8, 9, 10]:
                    ws.cell(row=row, column=c).border = thin_border
                
                # Label in col H
                cell_h = ws.cell(row=row, column=8, value=label)
                cell_h.font = Font(bold=True)
                cell_h.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formula in col J
                cell_j = ws.cell(row=row, column=10, value=f'=SUMIF(J4:J{total_row-1}, "{term}", I4:I{total_row-1})')
                cell_j.font = Font(bold=True)
                cell_j.alignment = Alignment(horizontal='center', vertical='center')
            
            os.makedirs(self.output_folder, exist_ok=True)
            wb.save(output_path)
            
            try:
                os.startfile(output_path)
            except Exception as e:
                pass # Ignore if it fails to open
                
            return True, f"Report successfully generated at:\n{output_path}"
            
        except Exception as e:
            err = traceback.format_exc()
            path_info = output_path if 'output_path' in locals() else 'Unknown'
            return False, f"Error processing: {str(e)}\n\nPath: {path_info}\n\n{err}"

class DestructionProcessor:
    def __init__(self, sap_file, destruction_file, output_folder, cost_center="1DMECD001", sloc="DHDD", plant="DM01", warehouse="Dammam Club"):
        self.sap_file = sap_file
        self.destruction_file = destruction_file
        self.output_folder = output_folder
        self.cost_center = cost_center or "1DMECD001"
        self.sloc = sloc or "DHDD"
        self.plant = plant or "DM01"
        self.warehouse = warehouse or "Dammam Club"

    def process(self):
        try:
            def load_with_header_detection(filepath):
                # First pass, read normally
                df = read_excel_fast(filepath)
                
                # Check if the header seems fake (lots of Unnamed columns)
                if sum("unnamed" in str(c).lower() for c in df.columns) >= len(df.columns) / 2:
                    # Scan the first 15 rows to find the real header
                    for idx, row in df.head(15).iterrows():
                        row_str = " ".join([str(x).lower() for x in row.values])
                        if "sku" in row_str or "material" in row_str or "qty" in row_str or "quantity" in row_str:
                            # Re-read with this row as the header. 
                            # idx is 0-indexed in the DataFrame, meaning it's row `idx + 1` in the file.
                            return read_excel_fast(filepath, header=idx + 1)
                return df

            sap_df = load_with_header_detection(self.sap_file)
            dest_df = load_with_header_detection(self.destruction_file)
            # Let's try to find a row that contains "SKU" or "Material" or looks like data
            def find_best_columns(df, is_sap=False):
                sku_col, name_col, qty_col, batch_col = None, None, None, None
                
                # Priority 1: Exact matches in headers
                for col in df.columns:
                    c_str = str(col).lower().strip()
                    if c_str == "material" or c_str == "sku" or c_str == "material code":
                        sku_col = col
                    elif c_str == "unrestricted" or c_str == "qty" or c_str == "quantity":
                        qty_col = col
                    elif c_str == "batch":
                        batch_col = col
                    elif "name" in c_str or "item" in c_str:
                        name_col = col

                # Priority 2: Data inspection if columns not found
                if not sku_col or not qty_col:
                    for col in df.columns:
                        # Inspect data (skip NaNs)
                        clean_data = df[col].dropna()
                        if clean_data.empty: continue
                        
                        sample = clean_data.astype(str)
                        
                        # Check for SKU pattern: mostly long alphanumeric strings
                        if any(len(s) > 8 and any(c.isdigit() for c in s) and any(c.isalpha() for c in s) for s in sample.head(10)):
                            if not sku_col: sku_col = col
                        
                        # Check for Quantity pattern: mostly numbers (ignoring headers)
                        # We use a more lenient numeric check
                        numeric_count = pd.to_numeric(clean_data, errors='coerce').notnull().sum()
                        if numeric_count > len(clean_data) * 0.3: # At least 30% are numbers
                            # If it's a date-like column, ignore it for quantity
                            if not any("-" in s or "/" in s or ":" in s for s in sample.head(5)):
                                if not qty_col or "unnamed" in str(qty_col).lower():
                                    qty_col = col
                        
                        # Check for Name pattern: longer descriptive strings
                        if name_col is None and sample.str.len().mean() > 10:
                            name_col = col
                            
                return sku_col, name_col, qty_col, batch_col

            sap_sku_col, _, sap_qty_col, sap_batch_col = find_best_columns(sap_df, is_sap=True)
            if not sap_batch_col: # Backup for batch
                for col in sap_df.columns:
                    if "batch" in str(col).lower():
                        sap_batch_col = col
                        break

            dest_sku_col, dest_name_col, dest_qty_col, _ = find_best_columns(dest_df)

            if not all([sap_sku_col, sap_batch_col, sap_qty_col]):
                return False, f"SAP file missing required columns (Material, Batch, Unrestricted). Found: {sap_df.columns.tolist()}"
            if not all([dest_sku_col, dest_qty_col]):
                return False, f"Destruction file missing SKU or Quantity columns. Found: {dest_df.columns.tolist()}"

            # Clean and convert
            sap_df[sap_sku_col] = sap_df[sap_sku_col].astype(str).str.strip()
            dest_df[dest_sku_col] = dest_df[dest_sku_col].astype(str).str.strip()
            
            # Ensure quantity columns are numeric
            sap_df[sap_qty_col] = pd.to_numeric(sap_df[sap_qty_col], errors='coerce').fillna(0)
            dest_df[dest_qty_col] = pd.to_numeric(dest_df[dest_qty_col], errors='coerce').fillna(0)
            
            # Group SAP by SKU
            sap_inventory = {}
            # Sort SAP by Batch to ensure consistent allocation (FIFO-ish)
            sap_df = sap_df.sort_values(by=[sap_sku_col, sap_batch_col])
            
            for sku, group in sap_df.groupby(sap_sku_col):
                batches = group[[sap_batch_col, sap_qty_col]].copy()
                sap_inventory[sku] = batches.to_dict('records')

            sap_id = ""
            try:
                sap_id = f"{os.path.basename(self.sap_file)}_{int(os.path.getmtime(self.sap_file))}_{os.path.getsize(self.sap_file)}"
            except:
                pass

            history_file = os.path.join(os.path.expanduser("~"), ".destruction_history_v2.json")
            history_data = {}
            if os.path.exists(history_file):
                try:
                    with open(history_file, 'r') as f:
                        history_data = json.load(f)
                except:
                    pass
            
            dest_file_id = os.path.basename(self.destruction_file)
            
            if sap_id in history_data:
                for old_dest_id, old_allocs in history_data[sap_id].items():
                    if old_dest_id == dest_file_id:
                        continue
                    for sku, batches_taken in old_allocs.items():
                        if sku in sap_inventory:
                            for batch_taken in batches_taken:
                                for inv_batch in sap_inventory[sku]:
                                    if str(inv_batch[sap_batch_col]) == batch_taken["batch"]:
                                        inv_batch[sap_qty_col] -= batch_taken["qty"]
                                        break
                                        
            current_allocs = {}

            results = []
            current_records = []
            for _, row in dest_df.iterrows():
                sku = str(row[dest_sku_col]).strip()
                # Skip header-like rows or empty rows
                if not sku or sku.lower() == "nan" or len(sku) < 5: # SKUs like FWDX... are long
                    continue
                    
                item_name = row[dest_name_col] if dest_name_col else "Unknown Item"
                try:
                    required_qty = float(row[dest_qty_col])
                except:
                    continue
                    
                if required_qty <= 0: continue
                
                original_required = required_qty
                
                # Check if this exact file, SKU, and quantity combination was already processed
                is_duplicate = False
                note_text = ""

                if sku not in sap_inventory:
                    results.append({
                        "Item Name": item_name,
                        "Material / SKU": sku,
                        "Allocated Qty": 0,
                        "Sloc": self.sloc,
                        "Cost Center": self.cost_center,
                        "Plant": self.plant,
                        "Batch": "NOT IN STOCK",
                        "Required Qty": original_required,
                        "Missing Qty": original_required,
                        "Note": note_text,
                        "Highlight": "Red"
                    })
                    continue

                batches = sap_inventory[sku]
                rows_added = []
                sku_allocs = []
                
                for batch_data in batches:
                    if required_qty <= 0:
                        break
                    
                    batch_id = str(batch_data[sap_batch_col])
                    available_qty = float(batch_data[sap_qty_col])
                    if available_qty <= 0:
                        continue

                    take_qty = min(required_qty, available_qty)
                    row_dict = {
                        "Item Name": item_name,
                        "Material / SKU": sku,
                        "Allocated Qty": take_qty,
                        "Sloc": self.sloc,
                        "Cost Center": self.cost_center,
                        "Plant": self.plant,
                        "Batch": batch_id,
                        "Required Qty": original_required,
                        "Missing Qty": 0,
                        "Note": note_text,
                        "Highlight": None
                    }
                    results.append(row_dict)
                    rows_added.append(row_dict)
                    sku_allocs.append({"batch": batch_id, "qty": take_qty})
                    
                    required_qty -= take_qty
                    batch_data[sap_qty_col] -= take_qty # Update in-memory inventory
                
                if sku_allocs:
                    if sku not in current_allocs:
                        current_allocs[sku] = []
                    current_allocs[sku].extend(sku_allocs)

                if required_qty > 0:
                    row_dict = {
                        "Item Name": item_name,
                        "Material / SKU": sku,
                        "Allocated Qty": 0,
                        "Sloc": self.sloc,
                        "Cost Center": self.cost_center,
                        "Plant": self.plant,
                        "Batch": "SHORTAGE",
                        "Required Qty": original_required,
                        "Missing Qty": required_qty,
                        "Note": "batch used and more qty not available in any batch",
                        "Highlight": "Red"
                    }
                    results.append(row_dict)
                    rows_added.append(row_dict)

                if len(rows_added) > 1:
                    for r in rows_added:
                        if r["Highlight"] is None:
                            r["Highlight"] = "Yellow"

            # Save the new list to history
            if sap_id not in history_data:
                history_data[sap_id] = {}
            history_data[sap_id][dest_file_id] = current_allocs
                
            try:
                with open(history_file, 'w') as f:
                    json.dump(history_data, f)
            except:
                pass

            if not results:
                return False, "No valid data found in destruction list. Ensure the SKUs and Quantities are in the correct columns."

            output_df = pd.DataFrame(results)
            highlight_flags = output_df["Highlight"].tolist()
            output_df = output_df.drop(columns=["Highlight"])
            
            today_str = datetime.datetime.now().strftime("%d-%m-%Y")
            wh_name = self.warehouse.strip() if hasattr(self, 'warehouse') and self.warehouse else ""
            if wh_name:
                base_filename = f"Destruction {wh_name} {today_str}"
            else:
                base_filename = f"Destruction {today_str}"
                
            output_path = get_unique_filename(self.output_folder, base_filename, "xlsx")
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name="Destruction")
                
            # Apply highlighting and formatting
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Destruction"]
            ws.sheet_view.showGridLines = False
            
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
            orange_fill = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for r_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[r_idx].height = 23
                highlight = highlight_flags[r_idx - 2] if r_idx > 1 else None
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=col)
                    cell.border = thin_border
                    
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.alignment = center_align
                    else:
                        if col in (1, 2):  # Item Name, SKU
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align
                            
                        if highlight == "Yellow":
                            cell.fill = yellow_fill
                        elif highlight == "Red":
                            cell.fill = red_fill
                        elif highlight == "Orange":
                            cell.fill = orange_fill
                            
            for col in ws.columns:
                letter = col[0].column_letter
                if letter == 'A':
                    ws.column_dimensions[letter].width = 35
                elif letter == 'B':
                    ws.column_dimensions[letter].width = 20
                else:
                    ws.column_dimensions[letter].width = 15
                        
            os.makedirs(self.output_folder, exist_ok=True)
            wb.save(output_path)
            
            try:
                os.startfile(output_path)
            except Exception as e:
                pass # Ignore if it fails to open
                
            return True, f"Destruction Report generated successfully at:\n{output_path}"
        except Exception as e:
            import traceback
            return False, f"Error in Destruction: {str(e)}\n\n{traceback.format_exc()}"






class TransferOrderProcessor:
    def __init__(self, to_file, master_file, output_folder, profile_settings):
        self.to_file = to_file
        self.master_file = master_file
        self.output_folder = output_folder
        
        # Apply defaults for empty fields
        self.profile = profile_settings or {}
        if not self.profile.get("branch"): self.profile["branch"] = "Dammam"
        if not self.profile.get("cost_center"): self.profile["cost_center"] = "1DMECD001"
        if not self.profile.get("sloc"): self.profile["sloc"] = "DHDD"
        if not self.profile.get("sloc2"): self.profile["sloc2"] = "DMHD"
        if not self.profile.get("plant"): self.profile["plant"] = "DM01"
        if not self.profile.get("warehouse"): self.profile["warehouse"] = "Dammam Club"
        if not self.profile.get("sloc_transfer"): self.profile["sloc_transfer"] = "DMHD"
        if not self.profile.get("signature"): self.profile["signature"] = "Rashid Saddique"
        
        self.warehouse = self.profile.get("warehouse", "Dammam Club").lower()
        
    def process(self):
        try:
            required = ["branch", "warehouse", "sloc", "plant", "sloc_transfer"]
            missing = [k for k in required if not self.profile.get(k)]
            if missing:
                return False, f"Missing required profile fields: {', '.join(missing)}. Please update your profile first."
                
            def load_with_fallback(f):
                if f.lower().endswith('.csv'): return pd.read_csv(f)
                else: return read_excel_fast(f)
                    
            master_df = load_with_fallback(self.master_file)
            sku_col, title_col = None, None
            for col in master_df.columns:
                c_str = str(col).lower()
                if "sku" in c_str: sku_col = col
                if "title" in c_str: title_col = col
                
            if not sku_col or not title_col:
                return False, "Master file must have columns containing 'Sku' and 'Title'"
                
            sku_to_title = {}
            for _, row in master_df.iterrows():
                sku = str(row[sku_col]).strip()
                title = str(row[title_col]).strip()
                if sku and sku != "nan":
                    sku_to_title[sku] = title
                    
            to_df = load_with_fallback(self.to_file)
            
            wh_col, to_sku_col, to_qty_col = None, None, None
            # Search for specific source warehouse naming first
            for col in to_df.columns:
                c_str = str(col).lower()
                if "source warehouse" in c_str or "source whse" in c_str:
                    wh_col = col
                    break
                    
            # Fallback to the 3rd column (index 2) if not found by name
            if not wh_col and len(to_df.columns) >= 3:
                wh_col = to_df.columns[2]
                
            for col in to_df.columns:
                c_str = str(col).lower()
                if not to_sku_col and "sku" in c_str: to_sku_col = col
                if not to_qty_col and ("qty" in c_str or "quantity" in c_str): to_qty_col = col
                
            if not wh_col or not to_sku_col or not to_qty_col:
                return False, "Transfer Order must have 'Source Warehouse', 'SKU', and 'Quantity' columns"
                
            # Clean and split the profile warehouse to find the key matching token (e.g. "dammam")
            ignore_words = {"club", "whse", "warehouse", "location", "hub", "station", "store"}
            profile_wh_words = [w for w in self.warehouse.lower().split() if w not in ignore_words]
            if not profile_wh_words:
                profile_wh_words = [self.warehouse.lower()]
                
            def matches_warehouse(val):
                v_str = str(val).lower()
                return any(w in v_str for w in profile_wh_words)
                
            filtered_df = to_df[to_df[wh_col].apply(matches_warehouse)]
                
            results = []
            for i, row in filtered_df.iterrows():
                sku = str(row[to_sku_col]).strip()
                qty = row[to_qty_col]
                
                if sku == "nan" or pd.isna(qty): continue
                try: qty = float(qty)
                except: continue
                
                title = sku_to_title.get(sku, "Unknown Item")
                
                results.append({
                    "S No.": len(results) + 1,
                    "Item Name": title,
                    "SKU": sku,
                    "Qty": qty,
                    "Sloc": self.profile.get("sloc"),
                    "Plant": self.profile.get("plant"),
                    "Sloc Transfer Pstg": self.profile.get("sloc_transfer")
                })
                
            if not results:
                available_wh = to_df[wh_col].dropna().unique().tolist()
                return False, f"No matching items found for your profile warehouse '{self.warehouse}'.\n\nThe Transfer Order file contains these warehouses:\n{', '.join(map(str, available_wh))}\n\nPlease update your profile to match."
                
            out_df = pd.DataFrame(results)
            today_str = datetime.datetime.now().strftime("%d-%m-%Y")
            wh_name = self.profile.get("warehouse", "").strip()
            if wh_name:
                base_filename = f"Destruction Sheet {wh_name} {today_str}"
            else:
                base_filename = f"Destruction Sheet {today_str}"
            out_path = get_unique_filename(self.output_folder, base_filename, "xlsx")
            
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                out_df.to_excel(writer, index=False, startrow=2, sheet_name="Destruction")
            
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = openpyxl.load_workbook(out_path)
            ws = wb["Destruction"]
            
            ws.sheet_view.showGridLines = False
            branch_name = self.profile.get("branch", "Unknown")
            ws.freeze_panes = 'A4'
            
            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            header1_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            header2_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header3_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws.merge_cells('A1:G1')
            ws['A1'] = f"{branch_name} Destruction"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = center_align
            ws['A1'].fill = header1_fill
            
            ws.merge_cells('A2:G2')
            ws['A2'] = f"Date: {datetime.datetime.now().strftime('%m/%d/%Y')}"
            ws['A2'].font = Font(size=12, bold=True)
            ws['A2'].alignment = center_align
            ws['A2'].fill = header2_fill
            
            N = ws.max_row
            empty_rows = 5
            total_row = N + empty_rows + 1
            
            max_col = ws.max_column
            for row in range(1, total_row + 2):
                ws.row_dimensions[row].height = 23
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if row <= total_row:
                        cell.border = thin_border
                    
                    # Left align Item Name (column 2)
                    if col == 2 and row > 3:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                        
                    if row == 3:
                        cell.fill = header3_fill
                        cell.font = Font(bold=True)
                    elif row > 3 and row <= total_row - 1:
                        cell.fill = white_fill
                        
            for c in range(1, max_col + 1):
                ws.cell(row=1, column=c).border = thin_border
                ws.cell(row=1, column=c).fill = header1_fill
                ws.cell(row=2, column=c).border = thin_border
                ws.cell(row=2, column=c).fill = header2_fill
            
            for col in ws.columns:
                max_length = 0
                column = col[2].column_letter
                for cell in col:
                    if cell.row >= 3:
                        try:
                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                        except: pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)
                
            ws.column_dimensions['C'].width = 30
            
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
            ws.cell(row=total_row, column=1, value="TOTAL")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(row=total_row, column=1).alignment = center_align
            
            qty_letter = openpyxl.utils.get_column_letter(4)
            ws.cell(row=total_row, column=4, value=f"=SUM({qty_letter}4:{qty_letter}{total_row-1})")
            
            pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(bold=True, color="9C0006")
            ws.cell(row=total_row, column=4).fill = pink_fill
            ws.cell(row=total_row, column=4).font = red_font
            ws.cell(row=total_row, column=4).alignment = center_align
            
            for c in range(1, max_col + 1):
                cell = ws.cell(row=total_row, column=c)
                cell.border = thin_border
                if c not in (1, 2, 3, 4):
                    cell.fill = white_fill
                
            if "dammam" in self.warehouse.lower() or "dammam" in branch_name.lower():
                sig_start_row = total_row + 4
                def add_sig(r, c_start, c_end, name):
                    ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
                    cell = ws.cell(row=r, column=c_start, value=name)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='bottom')
                    for c_idx in range(c_start, c_end + 1):
                        ws.cell(row=r, column=c_idx).border = Border(bottom=Side(style='thin', color='000000'))
                        
                add_sig(sig_start_row, 1, 3, "Mr. Emad Shahwan")
                add_sig(sig_start_row, 5, 7, "Mr.Rashid Siddique")
                
                sig_start_row_2 = sig_start_row + 4
                add_sig(sig_start_row_2, 1, 3, "Mr.Ahmed Safwat")
                add_sig(sig_start_row_2, 5, 7, "Mr.Ali")
                
            wb.save(out_path)
            try: os.startfile(out_path)
            except: pass
            return True, f"Transfer Order processed successfully!\nSaved to: {out_path}"
        except Exception as e:
            import traceback
            return False, f"Error processing Transfer Order: {str(e)}\n\n{traceback.format_exc()}"

class InvoiceProcessor:
    def __init__(self, po_file, master_file, output_folder):
        self.po_file = po_file
        self.master_file = master_file
        self.output_folder = output_folder
        
    def process(self):
        try:
            def load_with_fallback(f):
                if f.lower().endswith('.csv'): df = pd.read_csv(f)
                else: df = read_excel_fast(f)
                
                if sum("unnamed" in str(c).lower() for c in df.columns) >= len(df.columns) / 2:
                    for idx, row in df.head(15).iterrows():
                        row_str = " ".join([str(x).lower() for x in row.values])
                        if "desc" in row_str or "item" in row_str or "product" in row_str or "sku" in row_str or "title" in row_str:
                            if f.lower().endswith('.csv'): return pd.read_csv(f, header=idx + 1)
                            else: return read_excel_fast(f, header=idx + 1)
                return df
                
            def normalize_desc(d):
                return re.sub(r'\s+', ' ', str(d)).strip().lower()

            def clean_alphanumeric(d):
                return re.sub(r'[^a-z0-9]', '', normalize_desc(d))

            def is_desc_header(h):
                h_lower = str(h).lower()
                if any(x in h_lower for x in ["external id", "model number", "model num", "sku", "code"]):
                    return False
                return any(x in h_lower for x in ["title", "desc", "item", "product"])

            def is_qty_header(h):
                h_lower = str(h).lower()
                if "accepted" in h_lower or "received" in h_lower:
                    return False
                return any(x in h_lower for x in ["qty", "quantity"])

            def is_price_header(h):
                h_lower = str(h).lower()
                if "total" in h_lower:
                    return False
                return any(x in h_lower for x in ["price", "cost", "rate", "amount"])

            master_df = load_with_fallback(self.master_file)
            
            # Use EXACTLY the 1st and 2nd valid columns as Company Desc and PO Desc
            valid_cols = [c for c in master_df.columns if "unnamed" not in str(c).lower() or len(master_df[c].dropna()) > 0]
            
            comp_desc_col = valid_cols[0] if len(valid_cols) > 0 else master_df.columns[0]
            po_desc_col = valid_cols[1] if len(valid_cols) > 1 else master_df.columns[1]
            
            sku_col = master_df.columns[-1]
            for col in master_df.columns:
                if "sku" in str(col).lower() or "company code" in str(col).lower():
                    sku_col = col
            
            desc_to_data = {}
            for _, row in master_df.iterrows():
                m_comp = str(row[comp_desc_col]).strip()
                m_po = str(row[po_desc_col]).strip()
                sku = str(row[sku_col]).strip()
                
                # Match against both Company Desc and PO desc to find the item
                if m_comp and m_comp.lower() != "nan":
                    desc_to_data[normalize_desc(m_comp)] = {"sku": sku, "comp": m_comp, "po": m_po}
                if m_po and m_po.lower() != "nan":
                    desc_to_data[normalize_desc(m_po)] = {"sku": sku, "comp": m_comp, "po": m_po}

            ext = self.po_file.lower().split('.')[-1]
            po_data = []
            order_details = []
            
            if ext == "pdf":
                try: import pdfplumber
                except ImportError: return False, "pdfplumber is not installed."
                with pdfplumber.open(self.po_file) as pdf:
                    # Extract 'From' value if present in the text on page 0
                    from_val = ""
                    if len(pdf.pages) > 0:
                        first_page_text = pdf.pages[0].extract_text()
                        if first_page_text:
                            lines = [line.strip() for line in first_page_text.split('\n') if line.strip()]
                            for idx, line in enumerate(lines):
                                if line.lower() == "from" and idx + 1 < len(lines):
                                    from_val = lines[idx + 1]
                                    break
                    
                    for page_idx, page in enumerate(pdf.pages):
                        tables = page.extract_tables()
                        for table in tables:
                            # Check if this table is the items table or order details
                            is_items_table = False
                            header_idx = -1
                            for i, row in enumerate(table):
                                row_str = " ".join([str(c).lower() for c in row if c])
                                if "desc" in row_str or "item" in row_str or "title" in row_str:
                                    is_items_table = True
                                    header_idx = i
                                    break
                            
                            if is_items_table:
                                # Find column indices dynamically from the header row
                                header_row = table[header_idx]
                                desc_col_idx = -1
                                qty_col_idx = -1
                                price_col_idx = -1
                                
                                for idx, val in enumerate(header_row):
                                    if not val: continue
                                    v_str = str(val).lower().replace('\n', ' ').strip()
                                    if is_desc_header(v_str):
                                        desc_col_idx = idx
                                    elif "qty requested" in v_str or "quantity requested" in v_str or "qty. requested" in v_str:
                                        qty_col_idx = idx
                                    elif is_qty_header(v_str):
                                        if qty_col_idx == -1:
                                            qty_col_idx = idx
                                    elif is_price_header(v_str):
                                        price_col_idx = idx
                                
                                # Default fallbacks if not found
                                if desc_col_idx == -1:
                                    for idx, val in enumerate(header_row):
                                        if not val: continue
                                        if "title" in str(val).lower() or "desc" in str(val).lower():
                                            desc_col_idx = idx
                                            break
                                if desc_col_idx == -1: desc_col_idx = 0
                                
                                if qty_col_idx == -1:
                                    for idx, val in enumerate(header_row):
                                        if not val: continue
                                        if "qty" in str(val).lower() or "quantity" in str(val).lower():
                                            qty_col_idx = idx
                                            break
                                if qty_col_idx == -1: qty_col_idx = 1
                                
                                # Process rows after the header
                                for row in table[header_idx + 1:]:
                                    if len(row) <= desc_col_idx: continue
                                    desc_val = row[desc_col_idx]
                                    if not desc_val: continue
                                    
                                    desc = str(desc_val).strip()
                                    if desc.lower() in ["title", "description", "item", "external id"]:
                                        continue
                                    if desc.lower().startswith("total"):
                                        continue
                                        
                                    qty = "0"
                                    if qty_col_idx < len(row) and row[qty_col_idx]:
                                        qty = str(row[qty_col_idx]).replace(',', '').strip()
                                        
                                    price = "0"
                                    if price_col_idx >= 0 and price_col_idx < len(row) and row[price_col_idx]:
                                        price = str(row[price_col_idx]).replace(',', '').strip()
                                        
                                    po_data.append({"desc": desc, "qty": qty, "price": price})
                            else:
                                # Order Details table (first page, 2 columns, doesn't contain items)
                                if page_idx == 0 and len(table) > 0 and len(table[0]) == 2:
                                    order_details = [r for r in table if len(r) == 2 and r[0] is not None]
                                    if from_val:
                                        order_details.append(["From", from_val])
                                        
                    if not po_data:
                        # Fallback to regex text parsing if tables weren't found (e.g. Sharbatly PDFs without borders)
                        for page in pdf.pages:
                            text = page.extract_text()
                            if not text: continue
                            for line in text.split('\n'):
                                match = re.match(r'^(\d+\s+)?(.+?)\s+((?:[\d.,]+\s*){4,})$', line.strip(), re.IGNORECASE)
                                if match:
                                    full_name = match.group(2)
                                    numbers_str = match.group(3).strip()
                                    nums = re.split(r'\s+', numbers_str)
                                    
                                    last_word_match = re.search(r'([A-Za-z\u0600-\u06FF]+)([\d.,]+)$', full_name)
                                    if last_word_match:
                                        qty_from_name = last_word_match.group(2)
                                        full_name = full_name[:-len(qty_from_name)].strip()
                                        nums.insert(0, qty_from_name)
                                    
                                    qty_val = nums[0]
                                    price_val = nums[1]
                                    
                                    # Dynamically identify the "Total Amount Before VAT" column
                                    qty_f = 0.0
                                    price_f = 0.0
                                    try: qty_f = float(qty_val.replace(',', ''))
                                    except: pass
                                    try: price_f = float(price_val.replace(',', ''))
                                    except: pass
                                    
                                    gross = qty_f * price_f
                                    best_val = None
                                    
                                    for n_str in nums[2:]:
                                        try:
                                            n_f = float(n_str.replace(',', ''))
                                            if abs(n_f - gross) < 0.02:
                                                best_val = n_str
                                                break
                                        except: pass
                                        
                                    if best_val is None and len(nums) > 3:
                                        try:
                                            discount = float(nums[2].replace(',', ''))
                                            net = gross - discount
                                            for n_str in nums[3:]:
                                                try:
                                                    n_f = float(n_str.replace(',', ''))
                                                    if abs(n_f - net) < 0.02:
                                                        best_val = n_str
                                                        break
                                                except: pass
                                        except: pass
                                        
                                    if best_val is None:
                                        best_val = nums[3] if len(nums) >= 4 else nums[-1]
                                        
                                    value_val = best_val
                                    
                                    eng_name = re.sub(r'[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]+', '', full_name).strip()
                                    eng_name = re.sub(r'^\d+\s+', '', eng_name)
                                    eng_name = re.sub(r'\s+', ' ', eng_name)
                                    
                                    po_data.append({"desc": eng_name, "qty": qty_val, "price": price_val, "value": value_val})
            else:
                po_df = load_with_fallback(self.po_file)
                desc_col = None
                qty_col = None
                price_col = None
                
                # Check column headers dynamically
                for col in po_df.columns:
                    col_str = str(col).lower()
                    if is_desc_header(col_str):
                        desc_col = col
                    elif "qty requested" in col_str or "quantity requested" in col_str:
                        qty_col = col
                    elif is_qty_header(col_str):
                        if not qty_col:
                            qty_col = col
                    elif is_price_header(col_str):
                        price_col = col
                        
                # Looser fallbacks if not found by name
                if not desc_col:
                    desc_col = po_df.columns[0]
                if not qty_col and len(po_df.columns) > 1:
                    qty_col = po_df.columns[1]
                if not price_col and len(po_df.columns) > 2:
                    price_col = po_df.columns[2]
                
                for _, row in po_df.iterrows():
                    desc_val = str(row[desc_col]).strip() if pd.notnull(row[desc_col]) else ""
                    if not desc_val: continue
                    if desc_val.lower() in ["title", "description", "item", "external id"]:
                        continue
                    if desc_val.lower().startswith("total"):
                        continue
                        
                    q_val = str(row[qty_col]) if qty_col and pd.notnull(row[qty_col]) else "0"
                    p_val = str(row[price_col]) if price_col and pd.notnull(row[price_col]) else "0"
                    
                    po_data.append({
                        "desc": desc_val,
                        "qty": q_val,
                        "price": p_val
                    })
                        
            if not po_data:
                cols = ", ".join([str(c) for c in po_df.columns]) if 'po_df' in locals() else "None"
                return False, f"Could not extract item data from PO file.\nEnsure it has Description, Quantity, and Cost/Price columns.\nFound columns: {cols}"
                
            results = []
            for item in po_data:
                desc = item["desc"]
                if not desc or desc.lower() in ["nan", "none"]: continue
                
                norm_po_desc = normalize_desc(desc)
                if not norm_po_desc: continue
                
                sku = ""
                comp_desc = desc
                po_desc = desc
                
                # 1. Exact normalized match
                match = desc_to_data.get(norm_po_desc)
                
                # 2. Alphanumeric match fallback
                if not match:
                    po_alnum = clean_alphanumeric(norm_po_desc)
                    if po_alnum:
                        for m_desc, m_data in desc_to_data.items():
                            if clean_alphanumeric(m_desc) == po_alnum:
                                match = m_data
                                break
                                
                # 3. Substring match fallback
                if not match:
                    po_alnum = clean_alphanumeric(norm_po_desc)
                    if po_alnum:
                        for m_desc, m_data in desc_to_data.items():
                            m_alnum = clean_alphanumeric(m_desc)
                            if m_alnum and (po_alnum in m_alnum or m_alnum in po_alnum):
                                match = m_data
                                break
                                
                if match:
                    sku = match["sku"]
                    comp_desc = match["comp"]
                    po_desc = match["po"]
                            
                try: qty = float(re.sub(r'[^\d.]', '', item["qty"]))
                except: qty = 0
                try: price = float(re.sub(r'[^\d.]', '', item["price"]))
                except: price = 0
                
                results.append({
                    "Company Description": comp_desc,
                    "PO Description": po_desc,
                    "SKU / Company Code": sku,
                    "Quantity Requested": qty,
                    "Unit Cost": price
                })
                
            out_df = pd.DataFrame(results)
            
            today_str = datetime.datetime.now().strftime("%d-%m-%Y")
            source_base = os.path.splitext(os.path.basename(self.po_file))[0]
            base_filename = f"{source_base} {today_str}"
            out_path = get_unique_filename(self.output_folder, base_filename, "xlsx")
            
            # Start items table row 5 if order details exist, else row 1
            start_row = 4 if order_details else 0
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                out_df.to_excel(writer, index=False, startrow=start_row, sheet_name="Invoice PO")
            
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = openpyxl.load_workbook(out_path)
            ws = wb["Invoice PO"]
            
            ws.sheet_view.showGridLines = False
            
            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            details_hdr_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True)
            bold_font_12 = Font(size=12, bold=True)
            
            # 1. Format order details if present
            if order_details:
                for r_idx, row_data in enumerate(order_details, start=1):
                    ws.row_dimensions[r_idx].height = 23
                    key_cell = ws.cell(row=r_idx, column=1, value=row_data[0])
                    val_cell = ws.cell(row=r_idx, column=2, value=row_data[1])
                    
                    key_cell.font = bold_font
                    key_cell.fill = details_hdr_fill
                    key_cell.alignment = left_align
                    key_cell.border = thin_border
                    
                    val_cell.alignment = left_align
                    val_cell.fill = white_fill
                    val_cell.border = thin_border
                    
                    for col_idx in range(3, ws.max_column + 1):
                        ws.cell(row=r_idx, column=col_idx).fill = white_fill
                
                # Format blank spacer row 4
                ws.row_dimensions[4].height = 15
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=4, column=col_idx).fill = white_fill
                
                header_row = 5
            else:
                header_row = 1
            
            # 2. Format Items Table
            max_col = ws.max_column
            N = ws.max_row
            
            for row in range(header_row, N + 1):
                ws.row_dimensions[row].height = 23
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if row == header_row:
                        cell.font = bold_font_12
                        cell.fill = header_fill
                        cell.alignment = center_align
                    else:
                        cell.fill = white_fill
                        if col in (1, 2):
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align
            
            # 3. Add Total Row
            total_row = N + 2
            ws.row_dimensions[total_row].height = 25
            
            ws.row_dimensions[N+1].height = 15
            for col_idx in range(1, max_col + 1):
                ws.cell(row=N+1, column=col_idx).fill = white_fill
                
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
            tot_label = ws.cell(row=total_row, column=1, value="Total")
            tot_label.font = Font(size=12, bold=True)
            tot_label.alignment = center_align
            
            for col_idx in range(1, 4):
                ws.cell(row=total_row, column=col_idx).border = thin_border
                
            # Sum for Quantity Requested (column 4)
            qty_letter = openpyxl.utils.get_column_letter(4)
            qty_sum_formula = f"=SUM({qty_letter}{header_row+1}:{qty_letter}{N})"
            qty_sum_cell = ws.cell(row=total_row, column=4, value=qty_sum_formula)
            qty_sum_cell.font = Font(size=12, bold=True)
            qty_sum_cell.alignment = center_align
            qty_sum_cell.border = thin_border
            
            # Blank cell for Unit Cost (column 5)
            cost_cell = ws.cell(row=total_row, column=5)
            cost_cell.border = thin_border
            cost_cell.fill = white_fill
            
            # 4. Set Column Widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value and not str(cell.value).startswith('='):
                        try:
                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                        except: pass
                limit = 50 if column in ('A', 'B') else 25
                ws.column_dimensions[column].width = min(max_length + 2, limit)
                
            wb.save(out_path)
            
            try: os.startfile(out_path)
            except: pass
            
            return True, f"PO processed successfully!\nSaved to: {out_path}"
        except Exception as e:
            import traceback
            return False, f"Error processing PO: {str(e)}\n\n{traceback.format_exc()}"

class EcomInvoiceProcessor:
    def __init__(self, po_file, master_file, output_folder, dest_settings=None):
        self.po_file = po_file
        self.master_file = master_file
        self.output_folder = output_folder
        if dest_settings is None: dest_settings = {}
        self.plant = dest_settings.get("plant", "DM01")
        self.sloc2 = dest_settings.get("sloc2", "DMHD")
        self.branch = dest_settings.get("branch", "Dammam")
        self.signature = dest_settings.get("signature", "Rashid Saddique")

    def process(self):
        try:
            def load_with_fallback(f):
                if f.lower().endswith('.csv'):
                    return pd.read_csv(f)
                else:
                    return read_excel_fast(f)
                    
            master_df = load_with_fallback(self.master_file)
            
            if sum("unnamed" in str(c).lower() for c in master_df.columns) >= len(master_df.columns) / 2:
                for idx, row in master_df.head(15).iterrows():
                    row_str = " ".join([str(x).lower() for x in row.values])
                    if "desc" in row_str or "item" in row_str or "product" in row_str or "sku" in row_str or "title" in row_str:
                        master_df = load_with_fallback(self.master_file)
                        master_df.columns = master_df.iloc[idx]
                        master_df = master_df.iloc[idx+1:].reset_index(drop=True)
                        break

            desc_col = None
            amz_col = None
            code_col = None
            
            for col in master_df.columns:
                col_str = str(col).lower().strip()
                if "lab" in col_str:
                    desc_col = col
                elif "amazon" in col_str or "invoice" in col_str or "title" in col_str:
                    amz_col = col
                elif col_str == "sku" or "code" in col_str:
                    code_col = col

            # Fallbacks by index if not found
            valid_cols = [c for c in master_df.columns if "unnamed" not in str(c).lower() or len(master_df[c].dropna()) > 0]
            if not desc_col:
                desc_col = valid_cols[0] if len(valid_cols) > 0 else master_df.columns[0]
            if not amz_col:
                amz_col = valid_cols[1] if len(valid_cols) > 1 else (valid_cols[0] if len(valid_cols) > 0 else master_df.columns[0])
            if not code_col:
                code_col = valid_cols[2] if len(valid_cols) > 2 else (valid_cols[-1] if len(valid_cols) > 0 else master_df.columns[-1])

            master_records = []
            for _, row in master_df.iterrows():
                d_val = str(row[desc_col]).strip() if pd.notnull(row[desc_col]) else ""
                a_val = str(row[amz_col]).strip() if pd.notnull(row[amz_col]) else ""
                c_val = str(row[code_col]).strip() if pd.notnull(row[code_col]) else ""
                if d_val or a_val:
                    master_records.append({
                        "desc": d_val,
                        "amz": a_val,
                        "code": c_val
                    })

            ext = self.po_file.lower().split('.')[-1]
            po_data = []
            order_details = []
            
            noun_categories = {
                'apple': ['apple'],
                'banana': ['banana'],
                'beetroot': ['beet'],
                'beet': ['beet'],
                'blackberry': ['blackberry', 'blackberries'],
                'blueberry': ['blueberry', 'blueberries'],
                'onion': ['onion'],
                'tomato': ['tomato', 'tomatoes'],
                'pepper': ['pepper', 'peppers', 'capsicum', 'chilli', 'chili'],
                'chilli': ['chilli', 'chili', 'pepper'],
                'carrot': ['carrot', 'carrots'],
                'celery': ['celery'],
                'cucumber': ['cucumber'],
                'dragon': ['dragon'],
                'garlic': ['garlic'],
                'ginger': ['ginger'],
                'pineapple': ['pineapple'],
                'bean': ['bean', 'beans'],
                'lemon': ['lemon'],
                'lime': ['lime'],
                'lettuce': ['lettuce', 'iceberg', 'iceburge', 'romaine'],
                'iceberg': ['lettuce', 'iceberg', 'iceburge'],
                'iceburge': ['lettuce', 'iceberg', 'iceburge'],
                'orange': ['orange', 'easypeeler'],
                'nectarine': ['nectarine'],
                'plum': ['plum', 'plums'],
                'pomegranate': ['pomegranate'],
                'strawberry': ['strawberry', 'strawberries'],
                'walnut': ['walnut'],
                'grapes': ['grapes', 'grape'],
                'coconut': ['coconut'],
                'avocado': ['avocado'],
                'rucola': ['rucola', 'jarjeer', 'rocola', 'rocket', 'rocca'],
                'rocola': ['rucola', 'jarjeer', 'rocola', 'rocket', 'rocca'],
                'rocca': ['rucola', 'jarjeer', 'rocola', 'rocket', 'rocca'],
                'spinach': ['spinach'],
                'broccoli': ['broccoli'],
                'mushroom': ['mushroom', 'mushrooms'],
                'cabbage': ['cabbage'],
                'coriander': ['coriander'],
                'dill': ['dill'],
                'eggplant': ['eggplant', 'egg', 'plant'],
                'habak': ['habak', 'basil'],
                'basil': ['habak', 'basil'],
                'mint': ['mint'],
                'papaya': ['papaya'],
                'parsley': ['parsley', 'parsly'],
                'parsly': ['parsley', 'parsly'],
                'pear': ['pear'],
                'potato': ['potato', 'potatos'],
                'potatos': ['potato', 'potatos'],
                'zucchini': ['zucchini', 'koosa', 'zuccini', 'marrow'],
                'zuccini': ['zucchini', 'koosa', 'zuccini', 'marrow'],
                'koosa': ['zucchini', 'koosa', 'zuccini', 'marrow'],
                'mango': ['mango'],
                'pomelo': ['pomelo'],
                'peach': ['peach', 'peaches'],
                'almond': ['almond', 'aimond', 'almonds'],
                'flower': ['flower', 'flowers', 'tulip', 'tulips', 'rose', 'roses', 'carnation', 'carnations', 'matthiola', 'stem', 'stems', 'bouquet', 'chrysanthemum', 'lily', 'lilies'],
            }

            synonyms = {
                'rocola': ['rucola', 'rocket', 'jarjeer', 'arugula', 'leaf', 'leaves', 'jar', 'jir'],
                'rocca': ['rucola', 'jarjeer', 'arugula', 'leaf', 'leaves', 'jar', 'jir', 'rocket'],
                'potatos': ['potato'],
                'koosa': ['zucchini', 'marrow'],
                'parsly': ['parsley'],
                'taymor': ['timour'],
                'green': ['spring'],
                'spring': ['green'],
                'habak': ['basil'],
                'basil': ['habak'],
                'chilli': ['chili'],
                'iceburge': ['iceberg'],
                'bell': ['capsicum'],
                'capsicum': ['bell'],
                'smail': ['small', 'makdous'],
                'small': ['makdous'],
            }

            def stem_word(w):
                w = w.lower().strip()
                if w.endswith('ies') and len(w) > 4:
                    return w[:-3] + 'y'
                if w.endswith('es') and len(w) > 4:
                    if any(w.endswith(x) for x in ['ches', 'shes', 'xes', 'ses', 'oes']):
                        return w[:-2]
                    return w[:-1]
                if w.endswith('s') and not w.endswith('ss') and len(w) > 3:
                    return w[:-1]
                return w

            def split_camel_and_numbers(s):
                s = re.sub(r'([a-z])([A-Z])', r'\1 \2', str(s))
                s = re.sub(r'([A-Za-z])([0-9])', r'\1 \2', s)
                s = re.sub(r'([0-9])([A-Za-z])', r'\1 \2', s)
                return s

            def clean_text(t):
                t_split = split_camel_and_numbers(t)
                clean_str = re.sub(r'[^a-z0-9\s]', '', t_split.lower().strip())
                stemmed_words = [stem_word(w) for w in clean_str.split()]
                return " ".join(stemmed_words)

            def match_item(eng_name):
                eng_name_clean = clean_text(eng_name)
                tokens = set(eng_name_clean.split())
                
                expanded_tokens = set(tokens)
                for tok in tokens:
                    stemmed_tok = stem_word(tok)
                    syn_list = []
                    if stemmed_tok in synonyms:
                        syn_list.extend(synonyms[stemmed_tok])
                    if tok in synonyms:
                        syn_list.extend(synonyms[tok])
                    for s in syn_list:
                        expanded_tokens.add(stem_word(s))
                        
                pdf_categories = set()
                for tok in tokens:
                    for cat, aliases in noun_categories.items():
                        if tok in aliases:
                            pdf_categories.add(cat)
                            
                best_match = None
                best_score = -1.0
                
                for record in master_records:
                    desc_clean = clean_text(record['desc'])
                    amz_clean = clean_text(record['amz'])
                    
                    master_categories = set()
                    master_tokens = set(desc_clean.split() + amz_clean.split())
                    for tok in master_tokens:
                        for cat, aliases in noun_categories.items():
                            if tok in aliases:
                                master_categories.add(cat)
                                
                    if pdf_categories:
                        if master_categories and not master_categories.issubset(pdf_categories):
                            continue
                        match_found = False
                        for cat in pdf_categories:
                            if any(alias in desc_clean or alias in amz_clean for alias in noun_categories[cat]):
                                match_found = True
                                break
                        if not match_found:
                            continue
                    else:
                        stop_words = {'fresh', 'big', 'small', 'round', 'hot', 'yellow', 'green', 'red', 'white', 'dry', 'local', 'import', 'pcs', 'box', 'kg'}
                        meaningful_tokens = [tok for tok in tokens if tok not in stop_words]
                        if not meaningful_tokens:
                            continue
                        if not any(tok in desc_clean or tok in amz_clean for tok in meaningful_tokens):
                            continue
                    
                    intersection = expanded_tokens & master_tokens
                    if not intersection:
                        continue
                        
                    pdf_colors = {c for c in ['red', 'green', 'yellow', 'white'] if c in eng_name_clean}
                    if 'spring' in eng_name_clean:
                        pdf_colors.add('green')
                        
                    master_colors = set()
                    for c in ['red', 'green', 'yellow', 'white']:
                        if c in desc_clean or c in amz_clean:
                            master_colors.add(c)
                    if 'spring' in desc_clean or 'spring' in amz_clean:
                        master_colors.add('green')
                        
                    if pdf_colors and master_colors and not (pdf_colors & master_colors):
                        continue
                        
                    jaccard = len(intersection) / len(expanded_tokens | master_tokens)
                    
                    pdf_has_baby = 'baby' in eng_name_clean
                    master_has_baby = 'baby' in desc_clean or 'baby' in amz_clean
                    if master_has_baby and not pdf_has_baby:
                        jaccard -= 0.05
                    elif pdf_has_baby and master_has_baby:
                        jaccard += 0.05
                        
                    # Exact match bonus
                    if eng_name_clean == amz_clean or eng_name_clean == desc_clean:
                        jaccard += 0.50
                    elif eng_name_clean in amz_clean or eng_name_clean in desc_clean:
                        jaccard += 0.10
                        
                    if jaccard > best_score:
                        best_score = jaccard
                        best_match = record
                        
                if best_score >= 0.20:
                    return best_match
                return None

            if ext == "pdf":
                try:
                    import pdfplumber
                except ImportError:
                    return False, "pdfplumber is not installed."
                
                with pdfplumber.open(self.po_file) as pdf:
                    if len(pdf.pages) > 0:
                        first_page_text = pdf.pages[0].extract_text()
                        if first_page_text:
                            inv_no = ""
                            inv_date = ""
                            lines = first_page_text.split('\n')
                            for line in lines:
                                if "ﺓﺭﻭﺗﺎﻔﻟﺍ" in line or "رقم الفاتورة" in line or "invoice no" in line.lower() or "invoice number" in line.lower() or "inv no" in line.lower() or "invoice #" in line.lower():
                                    match_no = re.search(r'(\d+)', line)
                                    if match_no:
                                        inv_no = match_no.group(1)
                                if "date" in line.lower() or "ﺦـــﻳﺭﺎـﺗﻟﺍ" in line or "date:" in line.lower():
                                    match_date = re.search(r'(\d{2}[-/]\d{2}[-/]\d{4})', line)
                                    if match_date:
                                        inv_date = match_date.group(1)
                            if inv_no:
                                order_details.append(["رقم الفاتورة", inv_no])
                            if inv_date:
                                order_details.append(["التاريخ", inv_date])

                    for page in pdf.pages:
                        text = page.extract_text()
                        if not text: continue
                        for line in text.split('\n'):
                            # Match lines starting with a number and ending with at least 4 numbers (qty, price, discount, total...)
                            match = re.match(r'^(\d+\s+)?(.+?)\s+((?:[\d.,]+\s*){4,})$', line.strip(), re.IGNORECASE)
                            if match:
                                full_name = match.group(2)
                                numbers_str = match.group(3).strip()
                                nums = re.split(r'\s+', numbers_str)
                                
                                last_word_match = re.search(r'([A-Za-z\u0600-\u06FF]+)([\d.,]+)$', full_name)
                                if last_word_match:
                                    qty_from_name = last_word_match.group(2)
                                    full_name = full_name[:-len(qty_from_name)].strip()
                                    nums.insert(0, qty_from_name)
                                
                                if len(nums) < 2: continue
                                qty_val = nums[0]
                                price_val = nums[1]
                                
                                # Dynamically identify the "Total Amount Before VAT" column
                                qty_f = 0.0
                                price_f = 0.0
                                try: qty_f = float(qty_val.replace(',', ''))
                                except: pass
                                try: price_f = float(price_val.replace(',', ''))
                                except: pass
                                
                                gross = qty_f * price_f
                                best_val = None
                                
                                # Search for the exact gross value (Qty * Price)
                                for n_str in nums[2:]:
                                    try:
                                        n_f = float(n_str.replace(',', ''))
                                        if abs(n_f - gross) < 0.02:
                                            best_val = n_str
                                            break
                                    except: pass
                                    
                                # If gross not found, maybe there is a discount (assuming nums[2] is discount)
                                if best_val is None and len(nums) > 3:
                                    try:
                                        discount = float(nums[2].replace(',', ''))
                                        net = gross - discount
                                        for n_str in nums[3:]:
                                            try:
                                                n_f = float(n_str.replace(',', ''))
                                                if abs(n_f - net) < 0.02:
                                                    best_val = n_str
                                                    break
                                            except: pass
                                    except: pass
                                    
                                if best_val is None:
                                    # Fallback
                                    best_val = nums[3] if len(nums) >= 4 else nums[-1]
                                    
                                value_val = best_val
                                
                                eng_name = re.sub(r'[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]+', '', full_name).strip()
                                eng_name = re.sub(r'^\d+\s+', '', eng_name)
                                eng_name = re.sub(r'\s+', ' ', eng_name)
                                po_data.append({
                                    "desc": eng_name,
                                    "qty": qty_val,
                                    "price": price_val,
                                    "value": value_val
                                })
            else:
                po_df = load_with_fallback(self.po_file)
                if po_df is None or po_df.empty or len(po_df.columns) == 0:
                    return False, "Could not load Invoice file or file is empty."
                desc_col_po = None
                qty_col_po = None
                price_col_po = None
                
                for col in po_df.columns:
                    col_str = str(col).lower()
                    if is_desc_header(col_str):
                        desc_col_po = col
                    elif "qty" in col_str or "quantity" in col_str:
                        qty_col_po = col
                    elif is_price_header(col_str):
                        price_col_po = col
                
                if not desc_col_po:
                    desc_col_po = po_df.columns[0]
                if not qty_col_po and len(po_df.columns) > 1:
                    qty_col_po = po_df.columns[1]
                if not price_col_po and len(po_df.columns) > 2:
                    price_col_po = po_df.columns[2]
                
                for _, row in po_df.iterrows():
                    desc_val = str(row[desc_col_po]).strip() if pd.notnull(row[desc_col_po]) else ""
                    if not desc_val: continue
                    if desc_val.lower() in ["title", "description", "item", "external id"]:
                        continue
                    if desc_val.lower().startswith("total"):
                        continue
                    
                    q_val = str(row[qty_col_po]) if qty_col_po and pd.notnull(row[qty_col_po]) else "0"
                    p_val = str(row[price_col_po]) if price_col_po and pd.notnull(row[price_col_po]) else "0"
                    
                    val_col = None
                    for c in po_df.columns:
                        c_str = str(c).lower()
                        if "total" in c_str or "amount" in c_str or "value" in c_str or "قيمة" in c_str:
                            val_col = c
                            break
                    if not val_col:
                        if len(po_df.columns) > 3:
                            val_col = po_df.columns[3]
                        else:
                            val_col = po_df.columns[-1]
                    val_val = str(row[val_col]) if val_col and pd.notnull(row[val_col]) else "0"
                    
                    eng_name = re.sub(r'[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]+', '', desc_val).strip()
                    eng_name = re.sub(r'\s+', ' ', eng_name)
                    
                    po_data.append({
                        "desc": eng_name,
                        "qty": q_val,
                        "price": p_val,
                        "value": val_val
                    })

            if not po_data:
                return False, "Could not extract item data from Invoice/PO file."

            results = []
            for i, item in enumerate(po_data, 1):
                desc = item["desc"]
                if not desc or desc.lower() in ["nan", "none"]: continue
                
                match = match_item(desc)
                
                try: qty_amount = float(re.sub(r'[^\d.]', '', str(item.get("qty", "0"))))
                except: qty_amount = 0.0
                if qty_amount.is_integer():
                    qty_amount = int(qty_amount)
                    
                try: val_amount = float(re.sub(r'[^\d.]', '', str(item.get("value", "0"))))
                except: val_amount = 0.0
                
                unit_price = 0.0
                if qty_amount > 0:
                    unit_price = val_amount / qty_amount
                
                if match:
                    results.append({
                        "S.No": i,
                        "Product Name": match["desc"],
                        "Item Code": match["code"],
                        "Quantity": qty_amount,
                        "Total Amount": val_amount,
                        "Plant": self.plant,
                        "Storage Location": self.sloc2,
                        "Unit Price": unit_price,
                        "per pcs": "",
                        " ": "",
                        "  ": "",
                        "   ": "",
                        "    ": "",
                        "     ": "",
                        "      ": "",
                        "       ": "",
                        "        ": "",
                        "         ": "",
                        "          ": "",
                        "           ": "",
                        "            ": "",
                        "             ": ""
                    })
                else:
                    results.append({
                        "S.No": i,
                        "Product Name": "N/A",
                        "Item Code": "N/A",
                        "Quantity": qty_amount,
                        "Total Amount": val_amount,
                        "Plant": self.plant,
                        "Storage Location": self.sloc2,
                        "Unit Price": unit_price,
                        "per pcs": "",
                        " ": "",
                        "  ": "",
                        "   ": "",
                        "    ": "",
                        "     ": "",
                        "      ": "",
                        "       ": "",
                        "        ": "",
                        "         ": "",
                        "          ": "",
                        "           ": "",
                        "            ": "",
                        "             ": ""
                    })

            out_df = pd.DataFrame(results)
            
            today_str = datetime.datetime.now().strftime("%d-%m-%Y")
            source_base = os.path.splitext(os.path.basename(self.po_file))[0]
            base_filename = f"{source_base} {today_str}"
            out_path = get_unique_filename(self.output_folder, base_filename, "xlsx")
            
            start_row = 6
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                out_df.to_excel(writer, index=False, startrow=start_row, startcol=1, sheet_name="E-com Invoice")
                
            wb = openpyxl.load_workbook(out_path)
            ws = wb["E-com Invoice"]
            ws.sheet_view.showGridLines = False
            
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = 9
            
            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            details_hdr_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            center_align_no_wrap = Alignment(horizontal='center', vertical='center', wrap_text=False)
            center_align_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False)
            left_align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True)
            bold_font_12 = Font(size=12, bold=True)
            
            header_row = 7
            max_col = 10
            N = ws.max_row
            mismatch_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
            peach_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
                                    # Write Tally Sheet headers FIRST
            current_date = datetime.datetime.now().strftime("%B %d, %Y")
            
            # Row 2
            ws.cell(row=2, column=2, value="Storage:")
            ws.cell(row=2, column=4, value="E-Commerce Sharbatly.club")
            ws.cell(row=2, column=6, value="Date:")
            ws.cell(row=2, column=7, value=current_date)
            ws.cell(row=2, column=10, value="D/N No:")
            ws.cell(row=2, column=14, value="Transport")
            ws.cell(row=2, column=18, value="Start Time:")

            # Row 3
            ws.cell(row=3, column=2, value="Container No:")
            ws.cell(row=3, column=6, value="Arrival Port:")
            ws.cell(row=3, column=10, value="S. No:")
            ws.cell(row=3, column=14, value="Purchaser Name:")
            ws.cell(row=3, column=18, value="Ending Time:")
            
            # Row 4
            ws.cell(row=4, column=2, value="Origin:")
            ws.cell(row=4, column=4, value="Local: 20138")
            ws.cell(row=4, column=6, value="Temp Set:")
            ws.cell(row=4, column=10, value="Plup Tmp:")
            ws.cell(row=4, column=14, value="Truck Num")
            ws.cell(row=4, column=18, value="Plate No:")

            # Row 5 Title (with dynamic branch)
            title_cell = ws.cell(row=5, column=2, value=f"GOODS RECEIVED TALLY SHEET {self.branch.upper()} E-COM")
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            
            # Merging Headers
            ws.merge_cells("B5:F5") # Title merge B to F
            
            for r_idx in [2, 3, 4]:
                ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=3)   # B:C
                ws.merge_cells(start_row=r_idx, start_column=4, end_row=r_idx, end_column=5)   # D:E
                ws.merge_cells(start_row=r_idx, start_column=7, end_row=r_idx, end_column=9)   # G:I (Date/Arrival Port/Temp Set)
                ws.merge_cells(start_row=r_idx, start_column=11, end_row=r_idx, end_column=13) # K:M (D/N No, S. No, Plup Tmp data)
                    
                ws.merge_cells(start_row=r_idx, start_column=14, end_row=r_idx, end_column=15) # N:O
                ws.merge_cells(start_row=r_idx, start_column=16, end_row=r_idx, end_column=17) # P:Q
                ws.merge_cells(start_row=r_idx, start_column=18, end_row=r_idx, end_column=19) # R:S
                ws.merge_cells(start_row=r_idx, start_column=20, end_row=r_idx, end_column=21) # T:U

            # Format Tally headers (borders and bold)
            # Apply all borders on row 2, 3 and 4 but only invisible border line between the columns
            # specifically left border on F(6), J(10), O(15), S(19) and B(2), and right border on U(21)
            for r_idx in range(2, 5):
                for c_idx in range(2, 22): # B to U
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = bold_font_12
                    b_left = 'thin' if c_idx in [2, 6, 10, 14, 15, 18, 19] else None
                    b_right = 'thin' if c_idx == 21 else None
                    cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style=b_left), right=Side(style=b_right))

            # Row 5 Title (with dynamic branch)
            # Remove all borders from column G (7) onwards
            for c_idx in range(2, 22):
                cell = ws.cell(row=5, column=c_idx)
                if c_idx <= 6: # Columns B to F
                    cell.border = thin_border
                else:
                    cell.border = Border()

            # Ensure row 5 has black fill
            for col_idx in range(2, 7):
                ws.cell(row=5, column=col_idx).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

            # Row 6: left border on B
            for c_idx in range(2, 22):
                cell = ws.cell(row=6, column=c_idx)
                if c_idx == 2:
                    cell.border = Border(left=Side(border_style="thin", color="000000"))
                    
            # Header Row 7
            headers_map = {
                2: "S.No", 3: "Product Name", 5: "Item Code",
                8: "Quantity", 9: "Total Amount", 10: "Plant", 11: "Storage Location", 13: "Quantity", 14: "Unit Price"
            }
            for col in range(2, 15):
                cell = ws.cell(row=header_row, column=col)
                if col in headers_map:
                    cell.value = headers_map[col]
                cell.border = thin_border
                cell.font = bold_font_12
                cell.fill = header_fill
                cell.alignment = center_align_wrap
                
            ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=4) # C:D Product Name
            ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=7) # E:G Item Code
            ws.merge_cells(start_row=header_row, start_column=11, end_row=header_row, end_column=12) # K:L Storage Location

            light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            light_orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

            # Data Rows (8 to N)
            for row in range(header_row + 1, N + 1):
                ws.row_dimensions[row].height = 23
                
                # Reading data from original DataFrame output (starts at column B=2)
                sno = ws.cell(row=row, column=2).value
                prod = ws.cell(row=row, column=3).value
                code = ws.cell(row=row, column=4).value
                qty = ws.cell(row=row, column=5).value
                tot = ws.cell(row=row, column=6).value
                plnt = ws.cell(row=row, column=7).value
                sloc = ws.cell(row=row, column=8).value
                unit_price_val = ws.cell(row=row, column=9).value
                prc = ws.cell(row=row, column=10).value
                
                for c in range(1, 23): ws.cell(row=row, column=c).value = "" # Clear old positions
                
                ws.cell(row=row, column=2, value=sno)
                ws.cell(row=row, column=3, value=prod)
                ws.cell(row=row, column=5, value=code)
                ws.cell(row=row, column=8, value=qty)
                ws.cell(row=row, column=9, value=tot)
                ws.cell(row=row, column=10, value=plnt)
                ws.cell(row=row, column=11, value=sloc)
                ws.cell(row=row, column=13, value=f'=H{row}') # Quantity column mapping to H
                
                up_cell = ws.cell(row=row, column=14, value=f'=IFERROR(I{row}/H{row}, 0)')
                up_cell.number_format = '0.00'
                
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4) # C:D
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7) # E:G
                ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12) # K:L
                ws.merge_cells(start_row=row, start_column=18, end_row=row, end_column=20) # R:T
                
                is_mismatch = (prod == "N/A" or code == "N/A")
                
                for col in range(2, 22): # B to U
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if is_mismatch and col <= 15:
                        cell.fill = mismatch_fill
                    else:
                        if col == 8: # Quantity
                            cell.fill = light_green_fill
                        elif col == 9: # Total Amount
                            cell.fill = light_blue_fill
                        elif col == 14: # Unit Price
                            cell.fill = light_orange_fill
                        else:
                            cell.fill = white_fill
                            
                    if col == 3:
                        cell.alignment = left_align_no_wrap
                    else:
                        cell.alignment = center_align_no_wrap

            # Supervisor / Signature Row
            total_row = N + 1
            ws.row_dimensions[total_row].height = 25
            
            # Clear all values first and set thin borders for B to N (columns 2 to 14)
            for col_idx in range(2, 15):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.value = ""
                cell.border = thin_border
                cell.fill = white_fill
                
            # Merge C to G for supervisor details and signature
            ws.merge_cells(start_row=total_row, start_column=3, end_row=total_row, end_column=7) # C:G
            
            # Merge I:J
            ws.merge_cells(start_row=total_row, start_column=9, end_row=total_row, end_column=10) # I:J
            
            # Merge K:L
            ws.merge_cells(start_row=total_row, start_column=11, end_row=total_row, end_column=12) # K:L
            
            # Apply CellRichText with Calibri 11 Normal for prefix and Segoe Script 9 Bold for signature
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            sign_cell = ws.cell(row=total_row, column=3)
            sign_cell.value = CellRichText(
                TextBlock(font=InlineFont(rFont="Calibri", sz=11, b=False), text="E-commerce Supervisor: "),
                TextBlock(font=InlineFont(rFont="Segoe Script", sz=9, b=True), text=self.signature)
            )
            sign_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            
            # Re-apply borders on the merged and single cells (B to N) to render thin borders correctly
            for col_idx in range(2, 15):
                ws.cell(row=total_row, column=col_idx).border = thin_border
            
            # Calculate column widths
            col_widths = {
                'A': 8.5,
                'B': 8.5,    # 60px
                'C': 9.3,    # 65px
                'D': 28.5,   # 200px
                'E': 4.25,   # 30px (pixel-based conversion)
                'F': 12.75,  # 90px (pixel-based conversion)
                'G': 5.7,    # 40px (pixel-based conversion)
                'H': 9.3,    # 65px (pixel-based conversion)
                'I': 9.3,    # 65px (pixel-based conversion)
                'J': 9.3,    # 65px (pixel-based conversion)
                'K': 8.5,    # 60px
                'L': 8.5,    # 60px
                'M': 8.5,    # 60px
                'N': 8.5,    # 60px
                'O': 6.4,    # 45px
                'P': 11.4,   # 80px
                'Q': 9.3,    # 65px
                'R': 9.3,    # 65px
                'S': 9.3,    # 65px
                'T': 9.3,    # 65px
                'U': 9.3,    # 65px
                'V': 9.3     # 65px
            }
            for column, width in col_widths.items():
                ws.column_dimensions[column].width = width
                
            wb.save(out_path)
            
            try:
                os.startfile(out_path)
            except:
                pass
                
            return True, f"E-com Invoice processed successfully!\nSaved to: {out_path}"
        except Exception as e:
            import traceback
            return False, f"Error processing E-com Invoice: {str(e)}\n\n{traceback.format_exc()}"

import os
import threading


class RequestFormProcessor:
    def __init__(self, form_type, data, output_folder, dest_settings=None):
        self.form_type = form_type
        self.data = data
        self.output_folder = output_folder
        dest_settings = dest_settings or {}
        self.signature = dest_settings.get("signature", "Rashid Saddique")
        self.manager_name = dest_settings.get("manager_name", "Emad Shahwan.")
        self.manager_title = dest_settings.get("manager_title", "Dammam Sales Manager.")

    def process(self):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        import os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.form_type
        
        ws.sheet_view.showGridLines = False

        header_font = Font(size=12, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15

        ws['A1'] = f"TO: {self.manager_name}"
        ws['A2'] = f"{self.manager_title}"
        ws['A3'] = "Respected Sir."
        ws['D1'] = f"Date: {datetime.now().strftime('%d-%m-%Y')}"
        
        ws['A1'].font = header_font
        ws['A2'].font = header_font
        ws['A3'].font = header_font
        ws['D1'].font = header_font

        ws['A5'] = f"Subject: {self.form_type} Request"
        ws['A5'].font = header_font

        row_idx = 7
        
        header = self.data.get("header", {})
        rows = self.data.get("rows", [])
        
        if self.form_type == "Fuel":
            ws.cell(row=row_idx, column=1, value="Please allow the amount below as Fuel expenses to Montana Drivers.")
            row_idx += 2
            ws.cell(row=row_idx, column=1, value=f"DRIVER: {header.get('Driver Name', '')}").font = header_font
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=f"VAN / TRUCK #: {header.get('Van/Truck Number', '')}").font = header_font
            row_idx += 2
            
            headers = ["DATE", "KM", "AMOUNT"]
            amount_col_idx = 3
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Kilo Meters", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
                
        elif self.form_type == "Expenses":
            ws.cell(row=row_idx, column=1, value="Please allow the amounts below as Expenses.")
            row_idx += 2
            headers = ["DATE", "EXPENSE TYPE", "REMARKS", "AMOUNT"]
            amount_col_idx = 4
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Expense Type", ""), r.get("Remarks", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Invoices Attached for Reference")
            row_idx += 2
            
        elif self.form_type == "Loan":
            ws.cell(row=row_idx, column=1, value=f"Please allow the amounts below as Loan to {header.get('Employee Name', '')}.")
            row_idx += 2
            headers = ["DATE", "REASON", "AMOUNT"]
            amount_col_idx = 3
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Reason", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
            
        elif self.form_type == "Food Allowance":
            ws.cell(row=row_idx, column=1, value=f"Please allow below amounts as food allowance to Delivery Associate during trip to {header.get('Destination', '')}.")
            row_idx += 2
            headers = ["DATE", "AMOUNT"]
            amount_col_idx = 2
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
                
        elif self.form_type == "Transportation":
            ws.cell(row=row_idx, column=1, value=f"Please Allow below amounts as Transportation to {header.get('Employee Name', '')}.")
            row_idx += 2
            headers = ["DATE", "ROUTE / DETAILS", "AMOUNT"]
            amount_col_idx = 3
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Route/Details", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
                
        elif self.form_type == "Airport Parking":
            ws.cell(row=row_idx, column=1, value=f"Please Allow Bellow Amounts as Airport Parking Tickets for offloading to Customer: {header.get('Customer Name', '')}.")
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=f"Driver: {header.get('Driver Name', '')}").font = header_font
            row_idx += 2
            headers = ["DATE", "AMOUNT"]
            amount_col_idx = 2
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
                
        elif self.form_type == "Tires Fix":
            ws.cell(row=row_idx, column=1, value=f"Kindly allow below amounts for the tire Punctures/Fixes for Driver {header.get('Driver Name', '')}.")
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=f"VAN #: {header.get('Van Number', '')}").font = header_font
            row_idx += 2
            headers = ["DATE", "AMOUNT"]
            amount_col_idx = 2
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Invoices Attached for Reference")
            row_idx += 2
            
        elif self.form_type == "License Upgradation":
            ws.cell(row=row_idx, column=1, value=f"Please allow the amounts below for License Upgradation for {header.get('Employee/Van Name', '')}.")
            row_idx += 2
            headers = ["DATE", "REMARKS", "AMOUNT"]
            amount_col_idx = 3
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Remarks", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Invoices Attached for Reference")
            row_idx += 2
            
        elif self.form_type == "Club Supplies":
            ws.cell(row=row_idx, column=1, value="Please allow the amounts below to purchase Club Supplies needed for the team/branch.")
            row_idx += 2
            headers = ["DATE", "DETAILS", "AMOUNT"]
            amount_col_idx = 3
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            for r in rows:
                data_row = [r.get("Date", ""), r.get("Details", ""), r.get("Amount", "")]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Invoices Attached for Reference")
            row_idx += 2
            
        else: # Generic fallback / Custom Form
            intro = self.data.get("intro_text", f"Please allow the amounts below for {self.form_type} request.")
            ws.cell(row=row_idx, column=1, value=intro)
            row_idx += 2
            
            headers = self.data.get("columns", ["DATE", "AMOUNT"])
            headers = [h.upper() for h in headers]
            amount_col_idx = len(headers)
            
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            row_idx += 1
            
            col_keys = self.data.get("columns", ["Date", "Amount"])
            for r in rows:
                data_row = [r.get(key, "") for key in col_keys]
                for col, v in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col == amount_col_idx:
                        try:
                            cell.value = float(str(v).replace(',', ''))
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = v
                    else:
                        cell.value = v
                    cell.alignment = center_align
                row_idx += 1

        # Calculate Total and format the last row
        ws.cell(row=row_idx, column=1, value="TOTAL").font = header_font
        ws.cell(row=row_idx, column=1).border = thin_border
        
        # Merge columns between TOTAL and the AMOUNT column
        num_cols = len(headers)
        amount_col_idx = num_cols
        
        for c in range(2, amount_col_idx):
            ws.cell(row=row_idx, column=c).border = thin_border
            
        total_amount = 0
        for r in rows:
            try:
                total_amount += float(str(r.get("Amount", "0")).replace(',', ''))
            except:
                pass
                
        tot_cell = ws.cell(row=row_idx, column=amount_col_idx, value=total_amount)
        tot_cell.font = header_font
        tot_cell.border = thin_border
        tot_cell.number_format = '#,##0.00'
        tot_cell.alignment = center_align
        row_idx += 3

        # Signatures (Sales Manager left, Accounting middle, Supervisor right) - stuck within Columns A-D
        sm_cell = ws.cell(row=row_idx, column=1, value="Sales Manager")
        sm_cell.font = header_font
        sm_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Merge columns B & C for Accounting to perfectly center it between A and D
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        ac_cell = ws.cell(row=row_idx, column=2, value="Accounting")
        ac_cell.font = header_font
        ac_cell.alignment = Alignment(horizontal='center', vertical='center')

        sv_cell = ws.cell(row=row_idx, column=4, value="Supervisor")
        sv_cell.font = header_font
        sv_cell.alignment = Alignment(horizontal='right', vertical='center')
        row_idx += 3

        # Save
        import time
        safe_type = self.form_type.replace('/', '_').replace(' ', '_')
        base_filename = f"Request_{safe_type}_{int(time.time())}.xlsx"
        output_path = os.path.join(self.output_folder, base_filename)
        os.makedirs(self.output_folder, exist_ok=True)
        wb.save(output_path)
            
        return True, output_path
