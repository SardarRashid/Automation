import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

class ExcelGenerator:
    def __init__(self, export_path):
        self.export_path = export_path

    def get_file_path(self, month_year):
        if not os.path.exists(self.export_path):
            try:
                os.makedirs(self.export_path)
            except Exception as e:
                print(f"Error creating directory: {e}")
                
        return os.path.join(self.export_path, f"Sharbatly_Orders_{month_year}.xlsx")

    def _apply_headers(self, sheet, products):
        # General headers
        sheet.merge_cells('A1:Z1')
        sheet['A1'] = "WHOLE SALE MARKET SHEET"
        sheet['A1'].font = Font(bold=True, size=16, color="000080")
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet['A2'] = "DATE:"
        sheet['A2'].font = Font(bold=True)
        
        # Row 3 will have S.No, CUSTOMER, and then products
        sheet['A3'] = "S. No"
        sheet['B3'] = "CUSTOMER"
        
        header_font = Font(bold=True, size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        sheet['A3'].font = header_font
        sheet['A3'].alignment = header_align
        sheet['A3'].border = border_style
        
        sheet['B3'].font = header_font
        sheet['B3'].alignment = header_align
        sheet['B3'].border = border_style
        
        col_idx = 3
        for prod in products:
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = prod.get('name', 'Unknown')
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border_style
            sheet.column_dimensions[get_column_letter(col_idx)].width = 12
            col_idx += 1
            
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 25
        sheet.row_dimensions[3].height = 40

    def sync_orders(self, orders, products_dict):
        """
        orders: list of dicts from Firebase
        products_dict: dict of product {id: {name, ...}} from Firebase
        """
        if not self.export_path:
            return

        products_list = list(products_dict.values())
        product_ids = list(products_dict.keys())
        
        # Group orders by date
        orders_by_date = {}
        for order in orders:
            # Check if order is approved or delivered
            if order.get('status') in ['Cancelled', 'Draft', 'Rejected']:
                continue
                
            date_str = order.get('date', '') # YYYY-MM-DD
            if not date_str:
                continue
                
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            month_year = dt.strftime("%B_%Y") # e.g. June_2026
            day_sheet = dt.strftime("%d_%b") # e.g. 14_Jun
            
            if month_year not in orders_by_date:
                orders_by_date[month_year] = {}
            if day_sheet not in orders_by_date[month_year]:
                orders_by_date[month_year][day_sheet] = {"date": date_str, "orders": []}
                
            orders_by_date[month_year][day_sheet]["orders"].append(order)
            
        diag_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
            diagonal=Side(style='thin', color='000000'),
            diagonalUp=True
        )
        
        # Generate Excel files
        for month_year, days_data in orders_by_date.items():
            filepath = self.get_file_path(month_year)
            if os.path.exists(filepath):
                wb = load_workbook(filepath)
            else:
                wb = Workbook()
                wb.remove(wb.active) # Remove default sheet
                
            for day_sheet, data in days_data.items():
                if day_sheet in wb.sheetnames:
                    sheet = wb[day_sheet]
                    wb.remove(sheet)
                    
                sheet = wb.create_sheet(day_sheet)
                self._apply_headers(sheet, products_list)
                sheet['C2'] = data['date']
                
                # Write Orders
                row_idx = 4
                s_no = 1
                for order in data['orders']:
                    sheet.cell(row=row_idx, column=1, value=s_no).border = diag_border.copy(diagonal=None)
                    sheet.cell(row=row_idx, column=2, value=order.get('customerName', '')).border = diag_border.copy(diagonal=None)
                    
                    # Create a dictionary of order items for fast lookup
                    items = {item['productId']: item for item in order.get('items', [])}
                    
                    col_idx = 3
                    for p_id in product_ids:
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if p_id in items:
                            qty = items[p_id].get('qty', 0)
                            price = items[p_id].get('price', 0.0)
                            
                            # Format: Price top right, Qty bottom left
                            cell.value = f"       {price}\n{qty}"
                            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                            cell.border = diag_border
                            cell.font = Font(size=9, bold=True)
                        else:
                            cell.border = diag_border
                            
                        col_idx += 1
                        
                    sheet.row_dimensions[row_idx].height = 40
                    row_idx += 1
                    s_no += 1
                    
            # Update Items Sheet
            if "Master_Items" in wb.sheetnames:
                wb.remove(wb["Master_Items"])
            items_sheet = wb.create_sheet("Master_Items")
            items_sheet.append(["Product ID", "Code", "Name", "Category", "Price (SAR)", "Stock", "Unit"])
            for p_id, p_data in products_dict.items():
                items_sheet.append([
                    p_id,
                    p_data.get('code', ''),
                    p_data.get('name', ''),
                    p_data.get('category', ''),
                    p_data.get('price', 0),
                    p_data.get('stock', 0),
                    p_data.get('unit', '')
                ])
                
            # Formatting Items Sheet
            for cell in items_sheet[1]:
                cell.font = Font(bold=True)
                
            # Summary Sheet
            if "Summary" in wb.sheetnames:
                wb.remove(wb["Summary"])
            sum_sheet = wb.create_sheet("Summary", 0) # Place at start
            sum_sheet.append(["Date", "Total Orders", "Total Revenue (SAR)"])
            
            summary_data = []
            for d_sh, d_val in days_data.items():
                date_str = d_val['date']
                total_rev = sum(o.get('totalAmount', 0) for o in d_val['orders'])
                count = len(d_val['orders'])
                summary_data.append([date_str, count, total_rev])
                
            # Sort by date
            summary_data.sort(key=lambda x: x[0])
            for r in summary_data:
                sum_sheet.append(r)
                
            for cell in sum_sheet[1]:
                cell.font = Font(bold=True)
                
            wb.save(filepath)
            
        return True
